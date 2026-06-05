import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image
import os
import io
import bulk_deals_parser

def _safe_sheet_title(name, existing_titles):
    invalid_chars = '[]:*?/\\'
    cleaned = ''.join('_' if c in invalid_chars else c for c in str(name)).strip()
    if not cleaned:
        cleaned = "Sheet"
    base = cleaned[:28]
    title = base
    counter = 2
    while title in existing_titles:
        suffix = f"_{counter}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    existing_titles.add(title)
    return title

def export_to_excel(analyzer, file_path):
    if not analyzer.funds:
        return False

    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    
    # 1. Overview
    df_overview = analyzer.get_master_holdings(asset_type_filter=None)
    df_overview.to_excel(writer, sheet_name='Overview', index=False)
    
    # 2. Overlap Matrix
    overlap_matrices = analyzer.get_overlap_matrix()
    overlap_matrices['count'].to_excel(writer, sheet_name='Overlap Count')
    overlap_matrices['weight'].to_excel(writer, sheet_name='Overlap Weight')
    
    # 3. By Fund
    # Create a separate sheet for each fund or one merged sheet?
    # "one sub-table per fund" -> let's use one sheet with spacing or multiple sheets.
    # Instruction says "SHEET 3 'By Fund'", implying one sheet.
    wb = writer.book
    ws_by_fund = wb.create_sheet("By Fund")
    curr_row = 1
    for fund_name, info in analyzer.funds.items():
        ws_by_fund.cell(row=curr_row, column=1, value=f"Fund: {fund_name}")
        ws_by_fund.cell(row=curr_row, column=1).font = Font(bold=True, size=14)
        curr_row += 1
        
        fund_df = info['df']
        latest_date = info['dates'][-1]
        cols_to_show = ['Type', 'Name', 'Sector', latest_date]
        fund_data = fund_df[fund_df[latest_date] > 0][cols_to_show]
        
        for r in dataframe_to_rows(fund_data, index=False, header=True):
            for c_idx, value in enumerate(r, 1):
                ws_by_fund.cell(row=curr_row, column=c_idx, value=value)
            curr_row += 1
        curr_row += 2 # gap

    # 4. Concentration
    # Monthly allocation for top 20 most widely-held stocks
    df_overview_equity = analyzer.get_master_holdings(asset_type_filter='Equity')
    top_20 = df_overview_equity.sort_values(by='Funds Count', ascending=False).head(20)['Name'].tolist()
    ws_conc = wb.create_sheet("Concentration")
    conc_row = 1
    for stock in top_20:
        ws_conc.cell(row=conc_row, column=1, value=f"Stock: {stock}")
        ws_conc.cell(row=conc_row, column=1).font = Font(bold=True)
        conc_row += 1
        ts_df = analyzer.get_stock_time_series(stock)
        for r in dataframe_to_rows(ts_df, index=False, header=True):
            for c_idx, value in enumerate(r, 1):
                ws_conc.cell(row=conc_row, column=c_idx, value=value)
            conc_row += 1
        conc_row += 2

    # 5. Sector Breakdown
    df_sector = analyzer.get_sector_data()
    df_sector.to_excel(writer, sheet_name='Sector Breakdown')
    
    # 6. Trend - New/Exited
    ws_trend = wb.create_sheet("Trend - New-Exited")
    trends = analyzer.get_trends()
    tr_row = 1
    for fund, data in trends.items():
        ws_trend.cell(row=tr_row, column=1, value=fund).font = Font(bold=True)
        tr_row += 1
        
        ws_trend.cell(row=tr_row, column=1, value="New Entries")
        tr_row += 1
        for r in dataframe_to_rows(data['new_entries'], index=False, header=True):
            for c_idx, value in enumerate(r, 1):
                ws_trend.cell(row=tr_row, column=c_idx, value=value)
            tr_row += 1
        
        ws_trend.cell(row=tr_row, column=1, value="Exited")
        tr_row += 1
        for r in dataframe_to_rows(data['exited'], index=False, header=True):
            for c_idx, value in enumerate(r, 1):
                ws_trend.cell(row=tr_row, column=c_idx, value=value)
            tr_row += 1
        tr_row += 2

    # 7. Raw Data
    raw_list = []
    for fund_name, info in analyzer.funds.items():
        df = info['df'].copy()
        df['Fund'] = fund_name
        raw_list.append(df)
    if raw_list:
        pd.concat(raw_list).to_excel(writer, sheet_name='Raw Data', index=False)

    writer.close()
    
    # Post-processing with openpyxl for charts and formatting
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    
    # Overlap Matrix Conditional Formatting
    ws_overlap = wb['Overlap Weight']
    # Matrix starts at B2 (if df has index and header)
    max_r = ws_overlap.max_row
    max_c = ws_overlap.max_column
    if max_r > 1 and max_c > 1:
        color_rule = ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                                    end_type='num', end_value=100, end_color='FF0000')
        ws_overlap.conditional_formatting.add(f"B2:{get_column_letter(max_c)}{max_r}", color_rule)

    # Sector Breakdown Chart
    ws_sect = wb['Sector Breakdown']
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Sector Breakdown"
    chart.y_axis.title = '% Allocation'
    chart.x_axis.title = 'Sector'
    
    data = Reference(ws_sect, min_col=2, min_row=1, max_row=ws_sect.max_row, max_col=ws_sect.max_column-1)
    cats = Reference(ws_sect, min_col=1, min_row=2, max_row=ws_sect.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.grouping = "stacked"
    ws_sect.add_chart(chart, "H2")

    wb.save(file_path)
    return True

def export_prospects_report(analyzer, file_path, bulk_df=None, asset_type_filter="Equity"):
    prospects = analyzer.get_prospects_export_data(asset_type_filter=asset_type_filter)
    if not prospects:
        return False

    wb = Workbook()
    
    # 1. Summary Sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    headers = [
        "Rank", "Stock", "Sector", "Funds Holding", "Breadth %", 
        "Avg Momentum 3M", "Conviction %", "New Entrants", 
        "YF Status", "Current Price", "Active Buy Signal",
        "Price vs 52W High %", "Volume Spike", "Spike Date",
        "Price After Spike", "Bulk Deal Verdict", "Composite Score"
    ]
    ws_sum.append(headers)
    for cell in ws_sum[1]:
        cell.font = Font(bold=True)
    
    fill_green = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    fill_yellow = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    fill_red = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    fill_grey = PatternFill(start_color="e2e3e5", end_color="e2e3e5", fill_type="solid")

    for i, p in enumerate(prospects, 1):
        abs_sig = p['active_buy_signal']['signal'] if p['active_buy_signal'] else "N/A"
        yf_status = "Connected" if p.get('price_data_available') else "No Data"
        current_price = p['price_data']['current_price'] if p['price_data'] else "N/A"
        p52w = f"{p['price_data']['pct_below_52w_high']:.1f}%" if p['price_data'] else "N/A"
        if p['price_data']:
            volume_spike = "Yes" if p['price_data'].get('volume_spike') else "No"
            spike_date = p['price_data'].get('latest_volume_spike_date') or "N/A"
            post_spike = (
                f"{p['price_data']['price_change_since_volume_spike']:+.1f}%"
                if p['price_data'].get('price_change_since_volume_spike') is not None else "N/A"
            )
        else:
            volume_spike = "N/A"
            spike_date = "N/A"
            post_spike = "N/A"
        
        bulk_verdict = "N/A"
        if bulk_df is not None:
            activity = bulk_deals_parser.get_mf_bulk_activity(bulk_df, p['stock'])
            bulk_verdict = activity['verdict']

        row = [
            i, p['stock'], p['sector'], p['funds_holding'], f"{p['breadth_score']:.1f}%", 
            f"{p['momentum_score']:.2f}%", f"{p['conviction_score']:.1f}%", p['new_entrants'],
            yf_status, current_price, abs_sig, p52w, volume_spike, spike_date,
            post_spike, bulk_verdict, p['composite_score']
        ]
        ws_sum.append(row)
        
        fill = fill_red
        if p['composite_score'] >= 60: fill = fill_green
        elif p['composite_score'] >= 35: fill = fill_yellow
        
        for cell in ws_sum[ws_sum.max_row]:
            cell.fill = fill

    # Auto-width summary
    for col in ws_sum.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws_sum.column_dimensions[column].width = max_length + 2

    # 2-11. Individual Stock Sheets
    for p in prospects:
        s_name = _safe_sheet_title(p['stock'], set(wb.sheetnames))
        ws = wb.create_sheet(s_name)
        
        # Row 1: Title
        ws.merge_cells("A1:L1")
        cell_a1 = ws["A1"]
        cell_a1.value = p['stock']
        cell_a1.font = Font(bold=True, size=16)
        cell_a1.alignment = Alignment(horizontal="center")
        
        # Row 2: Metadata
        ws.append(["Sector", p['sector'], "Funds Holding", p['funds_holding'], "Composite Score", p['composite_score'], "Breadth %", f"{p['breadth_score']:.1f}%", "Momentum", f"{p['momentum_score']:.2f}%", "Conviction %", f"{p['conviction_score']:.1f}%"])
        
        # Monthly Allocation Table
        ws.append([])
        ws.append(["Monthly Allocation % - Actual Figures"])
        ws[ws.max_row][0].font = Font(bold=True)
        
        all_dates_set = set()
        for fd in p['fund_details']:
            all_dates_set.update(fd['monthly_series'].keys())
        all_dates = sorted(list(all_dates_set))
        
        header_row = ["Fund"] + [pd.to_datetime(d).strftime('%b-%y') for d in all_dates]
        ws.append(header_row)
        
        for fd in p['fund_details']:
            row = [fd['fund']]
            for d in all_dates:
                row.append(fd['monthly_series'].get(d, 0))
            ws.append(row)
            
            curr_row_idx = ws.max_row
            for j in range(2, len(header_row) + 1):
                val = ws.cell(row=curr_row_idx, column=j).value
                prev_val = ws.cell(row=curr_row_idx, column=j-1).value if j > 2 else None
                if prev_val is not None:
                    if val > prev_val: ws.cell(row=curr_row_idx, column=j).fill = fill_green
                    elif val < prev_val: ws.cell(row=curr_row_idx, column=j).fill = fill_red

        avg_row = ["AVERAGE"]
        for d in all_dates:
            vals = [fd['monthly_series'].get(d, 0) for fd in p['fund_details'] if fd['monthly_series'].get(d, 0) > 0]
            avg_row.append(sum(vals)/len(vals) if vals else 0)
        ws.append(avg_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        ws.append([])
        
        # Change Analysis Table
        ws.append(["Change Analysis"])
        ws[ws.max_row][0].font = Font(bold=True)
        ws.append(["Fund", "Latest %", "1M Change", "3M Change", "Trend"])
        
        for fd in p['fund_details']:
            trend = "Stable"
            if fd['change_3m'] > 0.5: trend = "Building"
            elif fd['change_3m'] < -0.5: trend = "Reducing"
            
            ws.append([fd['fund'], fd['latest_alloc'], f"{fd['change_1m']:+.2f}", f"{fd['change_3m']:+.2f}", trend])
            row_idx = ws.max_row
            if fd['change_1m'] > 0: ws.cell(row=row_idx, column=3).fill = fill_green
            elif fd['change_1m'] < 0: ws.cell(row=row_idx, column=3).fill = fill_red
            if fd['change_3m'] > 0: ws.cell(row=row_idx, column=4).fill = fill_green
            elif fd['change_3m'] < 0: ws.cell(row=row_idx, column=4).fill = fill_red

        ws.append([])

        # Bulk Deals Confirmation
        if bulk_df is not None:
            activity = bulk_deals_parser.get_mf_bulk_activity(bulk_df, p['stock'])
            ws.append(["Bulk Deals Confirmation"])
            ws[ws.max_row][0].font = Font(bold=True)
            ws.append(["Verdict", activity['verdict']])
            v_cell = ws.cell(row=ws.max_row, column=2)
            if activity['verdict'] == 'Strong MF Buying': v_cell.fill = fill_green
            elif activity['verdict'] == 'Mixed': v_cell.fill = fill_yellow
            elif activity['verdict'] == 'MF Selling': v_cell.fill = fill_red
            
            ws.append(["Date", "Client Name", "Buy/Sell", "Quantity", "Price", "Value"])
            for deal in activity['deals']:
                ws.append([deal['date'], deal['client_name'], deal['deal_type'], deal['quantity'], deal['price'], deal['quantity']*deal['price']])
            ws.append([])

        # Price Intelligence
        abs_info = p['active_buy_signal']
        if p['price_data_available']:
            pd_info = p['price_data']
            ws.append(["Price Intelligence"])
            ws[ws.max_row][0].font = Font(bold=True)
            ws.append(["Current Price", "52W High", "52W Low", "% Below 52W High"])
            ws.append([pd_info['current_price'], pd_info['price_52w_high'], pd_info['price_52w_low'], f"{pd_info['pct_below_52w_high']:.1f}%"])
            
            ws.append(["3M Price Change", "1M Price Change", "Active Buy Signal", "Explanation"])
            abs_sig = abs_info['signal'] if abs_info else "N/A"
            abs_expl = abs_info['explanation'] if abs_info else "N/A"
            ws.append([f"{pd_info['price_change_3m']:+.1f}%", f"{pd_info['price_change_1m']:+.1f}%", abs_sig, abs_expl])

            ws.append(["Volume Spike", "Spike Date", "Spike Ratio", "Price After Spike"])
            ws.append([
                "Yes" if pd_info.get('volume_spike') else "No",
                pd_info.get('latest_volume_spike_date') or "N/A",
                f"{pd_info['volume_spike_ratio']:.1f}x" if pd_info.get('volume_spike_ratio') is not None else "N/A",
                f"{pd_info['price_change_since_volume_spike']:+.1f}%"
                if pd_info.get('price_change_since_volume_spike') is not None else "N/A"
            ])
            
            abs_cell = ws.cell(row=ws.max_row, column=3)
            if abs_sig == 'strong_buy': abs_cell.fill = fill_green
            elif abs_sig == 'active_buy': abs_cell.fill = PatternFill(start_color="e9ffed", end_color="e9ffed", fill_type="solid")
            elif abs_sig == 'passive_drift': abs_cell.fill = fill_yellow
            elif abs_sig in ['partial_sell', 'reducing']: abs_cell.fill = fill_red
            
            ws.append([])

        # Score Breakdown
        ws.append(["Score Breakdown"])
        ws.append([f"Breadth: {p['breadth_score']:.1f} x 25% | Momentum: {p['momentum_score']*10:.1f} x 30% | Conviction: {p['conviction_score']:.1f} x 20% | Accel: {min(p['breadth_acceleration']*5, 15)} | Active Buy: {abs_info['score_bonus'] if abs_info else 0} | TOTAL: {p['composite_score']:.1f}"])
        
        img_data = io.BytesIO()
        p['figure'].savefig(img_data, format='png')
        img_data.seek(0)
        img = Image(img_data)
        ws.add_image(img, f"A{ws.max_row + 2}")

    wb.save(file_path)
    return True
